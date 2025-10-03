import os
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
from openpyxl import load_workbook
import threading
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing
import re
import sys

def is_dark_mode():
    """Detect if system is in dark mode (Windows only)"""
    if sys.platform == "win32":
        try:
            import winreg
            registry = winreg.ConnectRegistry(None, winreg.HKEY_CURRENT_USER)
            key = winreg.OpenKey(registry,
                                r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize")
            value, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
            winreg.CloseKey(key)
            return value == 0  # 0 = dark mode, 1 = light mode
        except:
            return False
    return False

def scan_single_file(file_path, targets_set, case_sensitive=False, use_regex=False):
    """Scan a single Excel file for target strings - optimized version"""
    try:
        # Use read_only=True for faster loading and lower memory usage
        wb = load_workbook(file_path, data_only=True, read_only=True)
        found_targets = set()

        for sheet in wb.sheetnames:
            # Early exit if all targets found
            if found_targets == targets_set:
                break

            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                # Early exit if all targets found
                if found_targets == targets_set:
                    break

                for cell in row:
                    if isinstance(cell, str):
                        for t in targets_set:
                            # Apply search options
                            match_found = False

                            if use_regex:
                                try:
                                    flags = 0 if case_sensitive else re.IGNORECASE
                                    if re.search(t, cell, flags):
                                        match_found = True
                                except re.error:
                                    # Invalid regex, fall back to literal search
                                    cell_cmp = cell if case_sensitive else cell.lower()
                                    t_cmp = t if case_sensitive else t.lower()
                                    if t_cmp in cell_cmp:
                                        match_found = True
                            else:
                                # Literal search
                                cell_cmp = cell if case_sensitive else cell.lower()
                                t_cmp = t if case_sensitive else t.lower()
                                if t_cmp in cell_cmp:
                                    match_found = True

                            if match_found:
                                found_targets.add(t)
                                # Early exit if all targets found
                                if found_targets == targets_set:
                                    break

        wb.close()

        if found_targets:
            return {"file": file_path, "found": list(found_targets)}
        return None

    except Exception as e:
        return {"file": file_path, "error": str(e)}

class TDScannerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("♡ TD Scanner v1.0 ♡")
        self.root.geometry("800x650")

        # Theme definitions
        self.themes = {
            "Clean Studio": {
                "bg_main": "#FAFAFA",
                "bg_secondary": "#FFFFFF",
                "bg_tertiary": "#F5F5F5",
                "text_color": "#1A1A1A",
                "accent": "#6366F1",
                "result_bg": "#FFFFFF",
                "result_fg": "#374151",
                "title": "TD Scanner",
                "font_title": ("Segoe UI", 28, "normal"),
                "font_main": ("Segoe UI", 11),
                "entry_bg": "#FFFFFF",
                "entry_fg": "#1A1A1A",
                "title_fg": "#6366F1",
                "btn_fg": "#FFFFFF",
                "label_targets": "Search Targets",
                "label_directory": "Scan Directory",
                "btn_browse": "Browse",
                "btn_scan": "Start Scan",
                "label_results": "SCAN RESULTS",
                "status_ready": "Ready to scan",
                "status_scanning": "Scanning... Please wait",
                "relief_style": "flat"
            },
            "Modern Y2K": {
                "bg_main": "#FFF0F5",
                "bg_secondary": "#FFB6E1",
                "bg_tertiary": "#FFFFFF",
                "text_color": "#2D2D2D",
                "accent": "#FF69B4",
                "result_bg": "#FFFFFF",
                "result_fg": "#FF1493",
                "title": "✨ TD Scanner ✨",
                "font_title": ("Segoe UI", 26, "bold"),
                "font_main": ("Segoe UI", 11),
                "entry_bg": "#FFFFFF",
                "entry_fg": "#2D2D2D",
                "title_fg": "#FF1493",
                "btn_fg": "#FFFFFF",
                "label_targets": "✨ Search Targets",
                "label_directory": "💕 Scan Directory",
                "btn_browse": "Browse",
                "btn_scan": "✨ Start Scan ✨",
                "label_results": "SCAN RESULTS",
                "status_ready": "Ready to scan!",
                "status_scanning": "Scanning... Please wait",
                "relief_style": "flat"
            },
            "Modern Minimal": {
                "bg_main": "#F8F9FA",
                "bg_secondary": "#E9ECEF",
                "bg_tertiary": "#FFFFFF",
                "text_color": "#212529",
                "accent": "#0D6EFD",
                "result_bg": "#FFFFFF",
                "result_fg": "#495057",
                "title": "TD Scanner",
                "font_title": ("Segoe UI", 24, "bold"),
                "font_main": ("Segoe UI", 11),
                "entry_bg": "#FFFFFF",
                "entry_fg": "#212529",
                "title_fg": "#0D6EFD",
                "btn_fg": "#FFFFFF",
                "label_targets": "Search Targets",
                "label_directory": "Scan Directory",
                "btn_browse": "Browse",
                "btn_scan": "Start Scan",
                "label_results": "SCAN RESULTS",
                "status_ready": "Ready to scan",
                "status_scanning": "Scanning...",
                "relief_style": "flat"
            },
            "Dark Modern": {
                "bg_main": "#1A1A1A",
                "bg_secondary": "#2D2D2D",
                "bg_tertiary": "#252525",
                "text_color": "#FFFFFF",
                "accent": "#00D9FF",
                "result_bg": "#0D0D0D",
                "result_fg": "#00FF9F",
                "title": "TD SCANNER",
                "font_title": ("Segoe UI", 24, "bold"),
                "font_main": ("Segoe UI", 11),
                "entry_bg": "#1A1A1A",
                "entry_fg": "#FFFFFF",
                "title_fg": "#00D9FF",
                "btn_fg": "#000000",
                "label_targets": "Search Targets",
                "label_directory": "Scan Directory",
                "btn_browse": "Browse",
                "btn_scan": "Start Scan",
                "label_results": "SCAN RESULTS",
                "status_ready": "Ready to scan",
                "status_scanning": "Scanning...",
                "relief_style": "flat"
            },
            "Lavender Dream": {
                "bg_main": "#E6E6FA",
                "bg_secondary": "#DDA0DD",
                "bg_tertiary": "#F5F5FF",
                "text_color": "#4B0082",
                "accent": "#9370DB",
                "result_bg": "#FFFFFF",
                "result_fg": "#6A5ACD",
                "title": "♡ TD Scanner ♡",
                "font_title": ("Segoe UI", 24, "bold"),
                "font_main": ("Segoe UI", 11),
                "entry_bg": "#FFFFFF",
                "entry_fg": "#4B0082",
                "title_fg": "#9370DB",
                "btn_fg": "#FFFFFF",
                "label_targets": "♡ Search Targets",
                "label_directory": "♡ Scan Directory",
                "btn_browse": "Browse",
                "btn_scan": "♡ Start Scan ♡",
                "label_results": "SCAN RESULTS",
                "status_ready": "Ready to scan ♡",
                "status_scanning": "Scanning...",
                "relief_style": "flat"
            },
            "Mint Fresh": {
                "bg_main": "#F0FFF4",
                "bg_secondary": "#B2F5EA",
                "bg_tertiary": "#FFFFFF",
                "text_color": "#234E52",
                "accent": "#38B2AC",
                "result_bg": "#FFFFFF",
                "result_fg": "#2C7A7B",
                "title": "🌿 TD Scanner",
                "font_title": ("Segoe UI", 24, "bold"),
                "font_main": ("Segoe UI", 11),
                "entry_bg": "#FFFFFF",
                "entry_fg": "#234E52",
                "title_fg": "#38B2AC",
                "btn_fg": "#FFFFFF",
                "label_targets": "🌿 Search Targets",
                "label_directory": "🌿 Scan Directory",
                "btn_browse": "Browse",
                "btn_scan": "Start Scan",
                "label_results": "SCAN RESULTS",
                "status_ready": "Ready to scan",
                "status_scanning": "Scanning...",
                "relief_style": "flat"
            },
            "Y2K Pink": {
                "bg_main": "#FF1493",
                "bg_secondary": "#00FFFF",
                "bg_tertiary": "#FFD700",
                "text_color": "#000000",
                "accent": "#7FFF00",
                "result_bg": "#000000",
                "result_fg": "#00FF00",
                "title": "☆★ TD SCANNER 2025 ★☆",
                "font_title": ("Comic Sans MS", 24, "bold"),
                "font_main": ("Comic Sans MS", 11),
                "entry_bg": "#FFFFFF",
                "entry_fg": "#000000",
                "title_fg": "#000000",
                "btn_fg": "#000000",
                "label_targets": "♪ Search Targets:",
                "label_directory": "♫ Scan Directory:",
                "btn_browse": "Browse ☺",
                "btn_scan": "♥ START SCAN ♥",
                "label_results": "◄ SCAN RESULTS ►",
                "status_ready": "● Ready to scan! ●",
                "status_scanning": "♪♫ Scanning... Please wait ♫♪",
                "relief_style": "raised"
            },
            "Cyber Purple": {
                "bg_main": "#9370DB",
                "bg_secondary": "#8A2BE2",
                "bg_tertiary": "#4B0082",
                "text_color": "#FFFFFF",
                "accent": "#FF00FF",
                "result_bg": "#000000",
                "result_fg": "#DA70D6",
                "title": "▲▼ TD SCANNER 2025 ▼▲",
                "font_title": ("Arial Black", 24, "bold"),
                "font_main": ("Arial", 11),
                "entry_bg": "#2D0052",
                "entry_fg": "#FFFFFF",
                "title_fg": "#FFFFFF",
                "btn_fg": "#000000",
                "label_targets": "◆ Search Targets:",
                "label_directory": "◆ Scan Directory:",
                "btn_browse": "Browse",
                "btn_scan": "▲ START SCAN ▼",
                "label_results": "◆ SCAN RESULTS ◇",
                "status_ready": "● Ready to scan! ●",
                "status_scanning": "◆ Scanning... ◇",
                "relief_style": "raised"
            },
            "Retro Green": {
                "bg_main": "#2E8B57",
                "bg_secondary": "#3CB371",
                "bg_tertiary": "#90EE90",
                "text_color": "#000000",
                "accent": "#00FF7F",
                "result_bg": "#001100",
                "result_fg": "#00FF00",
                "title": "░▒▓ TD SCANNER 2025 ▓▒░",
                "font_title": ("Courier New", 22, "bold"),
                "font_main": ("Courier New", 10),
                "entry_bg": "#C8FFC8",
                "entry_fg": "#000000",
                "title_fg": "#000000",
                "btn_fg": "#000000",
                "label_targets": "► Search Targets:",
                "label_directory": "► Scan Directory:",
                "btn_browse": "Browse ►",
                "btn_scan": "►► START SCAN ◄◄",
                "label_results": "▓▒░ SCAN RESULTS ░▒▓",
                "status_ready": "► Ready to scan ◄",
                "status_scanning": "►► Scanning... ◄◄",
                "relief_style": "ridge"
            },
            "Neon Blue": {
                "bg_main": "#1E90FF",
                "bg_secondary": "#00BFFF",
                "bg_tertiary": "#87CEEB",
                "text_color": "#000000",
                "accent": "#FFFF00",
                "result_bg": "#000033",
                "result_fg": "#00FFFF",
                "title": "◘◙ TD SCANNER 2025 ◙◘",
                "font_title": ("Impact", 24, "bold"),
                "font_main": ("Arial", 11),
                "entry_bg": "#E0FFFF",
                "entry_fg": "#000000",
                "title_fg": "#000000",
                "btn_fg": "#000000",
                "label_targets": "■ Search Targets:",
                "label_directory": "■ Scan Directory:",
                "btn_browse": "Browse",
                "btn_scan": "■ START SCAN ■",
                "label_results": "◘ SCAN RESULTS ◙",
                "status_ready": "■ Ready to scan ■",
                "status_scanning": "■ Scanning... ■",
                "relief_style": "raised"
            },
            "Sunset Orange": {
                "bg_main": "#FF6347",
                "bg_secondary": "#FF8C00",
                "bg_tertiary": "#FFD700",
                "text_color": "#000000",
                "accent": "#FF1493",
                "result_bg": "#2F1F1F",
                "result_fg": "#FFA500",
                "title": "✿❀ TD SCANNER 2025 ❀✿",
                "font_title": ("Georgia", 24, "bold"),
                "font_main": ("Georgia", 11),
                "entry_bg": "#FFF8DC",
                "entry_fg": "#000000",
                "title_fg": "#8B0000",
                "btn_fg": "#FFFFFF",
                "label_targets": "✿ Search Targets:",
                "label_directory": "✿ Scan Directory:",
                "btn_browse": "Browse",
                "btn_scan": "✿ START SCAN ✿",
                "label_results": "✦ SCAN RESULTS ✦",
                "status_ready": "✿ Ready to scan ✿",
                "status_scanning": "✦ Scanning... ✦",
                "relief_style": "raised"
            }
        }

        # Auto-detect system theme
        if is_dark_mode():
            self.current_theme = "Dark Modern"
        else:
            self.current_theme = "Clean Studio"

        self.target_entries = []  # List to store target entry widgets
        self.scan_cancelled = False
        self.executor = None
        self.case_sensitive_var = tk.BooleanVar(value=False)
        self.use_regex_var = tk.BooleanVar(value=False)
        self.apply_theme()
        self.create_widgets()

    def apply_theme(self):
        theme = self.themes[self.current_theme]
        self.bg_main = theme["bg_main"]
        self.bg_secondary = theme["bg_secondary"]
        self.bg_tertiary = theme["bg_tertiary"]
        self.text_color = theme["text_color"]
        self.accent = theme["accent"]
        self.result_bg = theme["result_bg"]
        self.result_fg = theme["result_fg"]
        self.title_text = theme["title"]
        self.font_title = theme["font_title"]
        self.font_main = theme["font_main"]
        self.entry_bg = theme["entry_bg"]
        self.entry_fg = theme["entry_fg"]
        self.title_fg = theme["title_fg"]
        self.btn_fg = theme["btn_fg"]
        self.label_targets = theme["label_targets"]
        self.label_directory = theme["label_directory"]
        self.btn_browse = theme["btn_browse"]
        self.btn_scan = theme["btn_scan"]
        self.label_results = theme["label_results"]
        self.status_ready = theme["status_ready"]
        self.status_scanning = theme["status_scanning"]
        self.relief_style = theme["relief_style"]
        self.root.configure(bg=self.bg_main)

    def change_theme(self, event=None):
        self.current_theme = self.theme_selector.get()
        self.apply_theme()
        self.refresh_ui()

    def refresh_ui(self):
        # Update all widget colors
        self.header.config(bg=self.bg_secondary, relief=self.relief_style)
        self.title_label.config(text=self.title_text, font=self.font_title,
                                bg=self.bg_secondary, fg=self.title_fg)

        self.theme_frame.config(bg=self.bg_main)
        self.theme_label.config(bg=self.bg_main, fg=self.text_color, font=self.font_main)

        self.controls.config(bg=self.bg_tertiary, relief=self.relief_style)
        self.target_main_frame.config(bg=self.bg_tertiary)
        self.targets_container.config(bg=self.bg_tertiary)
        self.targets_canvas.config(bg=self.bg_tertiary)
        self.target_label.config(text=self.label_targets, bg=self.bg_tertiary,
                                fg=self.text_color, font=self.font_main)
        self.add_target_btn.config(bg=self.accent, fg=self.btn_fg,
                                  font=self.font_main, relief=self.relief_style)

        # Update search options
        self.options_frame.config(bg=self.bg_tertiary)
        self.case_check.config(bg=self.bg_tertiary, fg=self.text_color,
                              font=self.font_main, selectcolor=self.entry_bg,
                              activebackground=self.bg_tertiary,
                              activeforeground=self.text_color)
        self.regex_check.config(bg=self.bg_tertiary, fg=self.text_color,
                               font=self.font_main, selectcolor=self.entry_bg,
                               activebackground=self.bg_tertiary,
                               activeforeground=self.text_color)

        # Update buttons
        self.btn_frame.config(bg=self.bg_tertiary)

        # Update all target entry frames
        for entry in self.target_entries:
            entry.master.config(bg=self.bg_tertiary)
            entry.config(bg=self.entry_bg, fg=self.entry_fg, font=self.font_main)
            for widget in entry.master.winfo_children():
                if isinstance(widget, tk.Label):
                    widget.config(bg=self.bg_tertiary, fg=self.text_color, font=self.font_main)

        self.dir_frame.config(bg=self.bg_tertiary)
        self.dir_label.config(text=self.label_directory, bg=self.bg_tertiary,
                             fg=self.text_color, font=self.font_main)
        self.dir_entry.config(bg=self.entry_bg, fg=self.entry_fg, font=self.font_main)
        self.browse_btn.config(text=self.btn_browse, bg=self.accent, fg=self.btn_fg,
                              font=self.font_main, relief=self.relief_style)

        self.scan_btn.config(text=self.btn_scan, bg=self.accent, fg=self.btn_fg,
                            font=self.font_main, relief=self.relief_style)

        self.progress_frame.config(bg=self.bg_tertiary)

        self.results_frame.config(bg=self.bg_main, relief=self.relief_style)
        self.results_label.config(text=self.label_results, bg=self.bg_main,
                                 fg=self.text_color, font=self.font_main)
        self.results_text.config(bg=self.result_bg, fg=self.result_fg)

        self.status_bar.config(text=self.status_ready, bg=self.bg_secondary,
                              fg=self.text_color, font=self.font_main)

    def create_widgets(self):
        # Theme Selector Frame
        self.theme_frame = tk.Frame(self.root, bg=self.bg_main)
        self.theme_frame.pack(fill="x", padx=10, pady=5)

        self.theme_label = tk.Label(self.theme_frame, text="Theme:",
                                    font=self.font_main,
                                    bg=self.bg_main, fg=self.text_color)
        self.theme_label.pack(side="left", padx=5)

        self.theme_selector = ttk.Combobox(self.theme_frame,
                                          values=list(self.themes.keys()),
                                          state="readonly", width=20,
                                          font=self.font_main)
        self.theme_selector.set(self.current_theme)
        self.theme_selector.bind("<<ComboboxSelected>>", self.change_theme)
        self.theme_selector.pack(side="left", padx=5)

        # Header Frame
        self.header = tk.Frame(self.root, bg=self.bg_secondary, bd=5, relief=self.relief_style)
        self.header.pack(fill="x", padx=10, pady=10)

        self.title_label = tk.Label(self.header, text=self.title_text,
                                    font=self.font_title,
                                    bg=self.bg_secondary, fg=self.title_fg)
        self.title_label.pack(pady=10)

        # Controls Frame
        self.controls = tk.Frame(self.root, bg=self.bg_tertiary, bd=5, relief=self.relief_style)
        self.controls.pack(fill="x", padx=10, pady=5)

        # Search Target Input
        self.target_main_frame = tk.Frame(self.controls, bg=self.bg_tertiary)
        self.target_main_frame.pack(pady=10, fill="x")

        # Header with label and add button
        target_header = tk.Frame(self.target_main_frame, bg=self.bg_tertiary)
        target_header.pack(fill="x", padx=5)

        self.target_label = tk.Label(target_header, text=self.label_targets,
                                     font=self.font_main, bg=self.bg_tertiary,
                                     fg=self.text_color)
        self.target_label.pack(side="left", padx=5)

        self.add_target_btn = tk.Button(target_header, text="+ Add Target",
                                        command=self.add_target_field,
                                        font=self.font_main, bg=self.accent,
                                        fg=self.btn_fg, bd=2, relief=self.relief_style,
                                        cursor="hand2")
        self.add_target_btn.pack(side="left", padx=5)

        # Scrollable container for target entries
        scroll_frame = tk.Frame(self.target_main_frame, bg=self.bg_tertiary)
        scroll_frame.pack(fill="both", padx=5, pady=5, expand=False)

        # Canvas for scrolling
        self.targets_canvas = tk.Canvas(scroll_frame, bg=self.bg_tertiary,
                                        height=150, highlightthickness=0)
        self.targets_scrollbar = tk.Scrollbar(scroll_frame, orient="vertical",
                                             command=self.targets_canvas.yview)

        # Frame inside canvas
        self.targets_container = tk.Frame(self.targets_canvas, bg=self.bg_tertiary)

        # Configure canvas
        self.targets_canvas.configure(yscrollcommand=self.targets_scrollbar.set)

        # Pack scrollbar and canvas
        self.targets_scrollbar.pack(side="right", fill="y")
        self.targets_canvas.pack(side="left", fill="both", expand=True)

        # Create window in canvas
        self.canvas_frame = self.targets_canvas.create_window((0, 0),
                                                              window=self.targets_container,
                                                              anchor="nw")

        # Bind configuration
        self.targets_container.bind("<Configure>", self.on_targets_configure)
        self.targets_canvas.bind("<Configure>", self.on_canvas_configure)

        # Bind mousewheel for scrolling
        self.targets_canvas.bind_all("<MouseWheel>", self.on_mousewheel)

        # Add first target field
        self.add_target_field("orgEmpCertDetail")

        # Directory Selection
        self.dir_frame = tk.Frame(self.controls, bg=self.bg_tertiary)
        self.dir_frame.pack(pady=5)

        self.dir_label = tk.Label(self.dir_frame, text=self.label_directory,
                                 font=self.font_main, bg=self.bg_tertiary,
                                 fg=self.text_color)
        self.dir_label.pack(side="left", padx=5)

        self.dir_entry = tk.Entry(self.dir_frame, width=40, font=self.font_main,
                                  bg=self.entry_bg, fg=self.entry_fg, bd=3)
        self.dir_entry.insert(0, r"D:\DreamSVN\Dream_Doc\1.ProgramSpec\X.Version\3.0")
        self.dir_entry.pack(side="left", padx=5)

        self.browse_btn = tk.Button(self.dir_frame, text=self.btn_browse,
                                    command=self.browse_directory,
                                    font=self.font_main, bg=self.accent,
                                    fg=self.btn_fg, bd=3, relief=self.relief_style,
                                    cursor="hand2")
        self.browse_btn.pack(side="left", padx=5)

        # Search Options
        self.options_frame = tk.Frame(self.controls, bg=self.bg_tertiary)
        self.options_frame.pack(pady=10)

        self.case_check = tk.Checkbutton(self.options_frame, text="Case Sensitive",
                                        variable=self.case_sensitive_var,
                                        bg=self.bg_tertiary, fg=self.text_color,
                                        font=self.font_main, selectcolor=self.entry_bg,
                                        activebackground=self.bg_tertiary,
                                        activeforeground=self.text_color)
        self.case_check.pack(side="left", padx=10)

        self.regex_check = tk.Checkbutton(self.options_frame, text="Use Regex",
                                         variable=self.use_regex_var,
                                         bg=self.bg_tertiary, fg=self.text_color,
                                         font=self.font_main, selectcolor=self.entry_bg,
                                         activebackground=self.bg_tertiary,
                                         activeforeground=self.text_color)
        self.regex_check.pack(side="left", padx=10)

        # Scan and Cancel Buttons
        self.btn_frame = tk.Frame(self.controls, bg=self.bg_tertiary)
        self.btn_frame.pack(pady=10)

        self.scan_btn = tk.Button(self.btn_frame, text=self.btn_scan,
                                 command=self.start_scan,
                                 font=self.font_main, bg=self.accent,
                                 fg=self.btn_fg, bd=5, relief=self.relief_style,
                                 cursor="hand2", width=20)
        self.scan_btn.pack(side="left", padx=5)

        self.cancel_btn = tk.Button(self.btn_frame, text="Cancel",
                                    command=self.cancel_scan,
                                    font=self.font_main, bg="#FF6347",
                                    fg="white", bd=5, relief=self.relief_style,
                                    cursor="hand2", width=10, state="disabled")
        self.cancel_btn.pack(side="left", padx=5)

        # Progress Bar
        self.progress_frame = tk.Frame(self.controls, bg=self.bg_tertiary)
        self.progress_frame.pack(pady=5, fill="x", padx=20)

        self.progress_bar = ttk.Progressbar(self.progress_frame, mode='determinate',
                                           length=400)
        self.progress_bar.pack(fill="x", expand=True)

        # Results Frame
        self.results_frame = tk.Frame(self.root, bg=self.bg_main, bd=5, relief=self.relief_style)
        self.results_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.results_label = tk.Label(self.results_frame, text=self.label_results,
                                     font=self.font_main, bg=self.bg_main,
                                     fg=self.text_color)
        self.results_label.pack(pady=5)

        self.results_text = scrolledtext.ScrolledText(self.results_frame,
                                                     font=("Courier New", 10),
                                                     bg=self.result_bg, fg=self.result_fg,
                                                     bd=3, relief="sunken")
        self.results_text.pack(fill="both", expand=True, padx=10, pady=10)

        # Status Bar
        self.status_bar = tk.Label(self.root, text=self.status_ready,
                                  font=self.font_main, bg=self.bg_secondary,
                                  fg=self.text_color, bd=3, relief="sunken")
        self.status_bar.pack(fill="x", side="bottom")

    def on_targets_configure(self, event=None):
        # Update scrollregion when targets container size changes
        self.targets_canvas.configure(scrollregion=self.targets_canvas.bbox("all"))

    def on_canvas_configure(self, event):
        # Update the width of the canvas window to match canvas width
        canvas_width = event.width
        self.targets_canvas.itemconfig(self.canvas_frame, width=canvas_width)

    def on_mousewheel(self, event):
        # Enable mousewheel scrolling
        self.targets_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def add_target_field(self, default_value=""):
        # Create a frame for this target entry
        entry_frame = tk.Frame(self.targets_container, bg=self.bg_tertiary)
        entry_frame.pack(fill="x", pady=3)

        # Target number label
        target_num = len(self.target_entries) + 1
        num_label = tk.Label(entry_frame, text=f"#{target_num}:",
                            font=self.font_main, bg=self.bg_tertiary,
                            fg=self.text_color, width=3)
        num_label.pack(side="left", padx=2)

        # Entry field
        entry = tk.Entry(entry_frame, width=40, font=self.font_main,
                        bg=self.entry_bg, fg=self.entry_fg, bd=3)
        if default_value:
            entry.insert(0, default_value)
        entry.pack(side="left", padx=5)

        # Remove button
        remove_btn = tk.Button(entry_frame, text="✖",
                              command=lambda: self.remove_target_field(entry_frame, entry),
                              font=self.font_main, bg="#FF6347",
                              fg="white", bd=2, relief=self.relief_style,
                              cursor="hand2", width=3)
        remove_btn.pack(side="left", padx=2)

        # Store reference
        self.target_entries.append(entry)

        # Update scroll region
        self.targets_container.update_idletasks()
        self.targets_canvas.configure(scrollregion=self.targets_canvas.bbox("all"))

    def remove_target_field(self, frame, entry):
        if len(self.target_entries) <= 1:
            messagebox.showwarning("Warning", "You need at least one search target!")
            return

        self.target_entries.remove(entry)
        frame.destroy()
        self.update_target_numbers()

    def update_target_numbers(self):
        # Update the numbering after removal
        for i, entry in enumerate(self.target_entries, 1):
            parent = entry.master
            for widget in parent.winfo_children():
                if isinstance(widget, tk.Label) and widget.cget("text").startswith("#"):
                    widget.config(text=f"#{i}:")
                    break

    def browse_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.dir_entry.delete(0, tk.END)
            self.dir_entry.insert(0, directory)

    def start_scan(self):
        root_dir = self.dir_entry.get().strip()

        # Collect all target values from entry fields
        targets = []
        for entry in self.target_entries:
            target = entry.get().strip()
            if target:
                targets.append(target)

        if not targets:
            messagebox.showerror("Error!", "Please enter at least one search target!")
            return

        if not os.path.exists(root_dir):
            messagebox.showerror("Error!", "Directory does not exist!")
            return

        # Get search options
        case_sensitive = self.case_sensitive_var.get()
        use_regex = self.use_regex_var.get()

        # Reset cancellation flag
        self.scan_cancelled = False

        # Update UI
        self.results_text.delete(1.0, tk.END)
        self.status_bar.config(text=self.status_scanning, bg=self.accent)
        self.progress_bar["value"] = 0
        self.scan_btn.config(state="disabled")
        self.cancel_btn.config(state="normal")

        # Run scan in separate thread
        thread = threading.Thread(target=self.perform_scan,
                                 args=(targets, root_dir, case_sensitive, use_regex))
        thread.daemon = True
        thread.start()

    def cancel_scan(self):
        """Cancel the ongoing scan"""
        self.scan_cancelled = True
        if self.executor:
            self.executor.shutdown(wait=False, cancel_futures=True)
        self.update_status("Scan cancelled by user", self.bg_secondary)
        self.scan_btn.config(state="normal")
        self.cancel_btn.config(state="disabled")

    def perform_scan(self, targets, root_dir, case_sensitive=False, use_regex=False):
        targets_set = set(targets)
        results = []

        try:
            # First, collect all Excel file paths
            file_paths = []
            for dirpath, dirnames, filenames in os.walk(root_dir):
                for file in filenames:
                    if "TD" in file and file.endswith(".xlsx") and not file.startswith("~$"):
                        file_paths.append(os.path.join(dirpath, file))

            file_count = len(file_paths)
            self.update_status(f"Found {file_count} files to scan...")

            if file_count == 0:
                self.display_results([], 0)
                self.scan_btn.config(state="normal")
                self.cancel_btn.config(state="disabled")
                return

            # Use multiprocessing to scan files in parallel
            # Use max 4 processes to avoid overwhelming the system
            max_workers = min(4, multiprocessing.cpu_count())
            processed = 0

            with ProcessPoolExecutor(max_workers=max_workers) as executor:
                self.executor = executor

                # Submit all tasks
                future_to_file = {
                    executor.submit(scan_single_file, fp, targets_set,
                                  case_sensitive, use_regex): fp
                    for fp in file_paths
                }

                # Process completed tasks
                for future in as_completed(future_to_file):
                    # Check if scan was cancelled
                    if self.scan_cancelled:
                        break

                    processed += 1
                    file_path = future_to_file[future]
                    file_name = os.path.basename(file_path)

                    # Update progress
                    progress = (processed / file_count) * 100
                    self.root.after(0, lambda p=progress: self.progress_bar.config(value=p))
                    self.update_status(f"Progress: {processed}/{file_count} - {file_name}")

                    try:
                        result = future.result()
                        if result:
                            if "error" in result:
                                self.append_result(f"❌ Error: {result['file']}\n   {result['error']}\n\n")
                            else:
                                results.append(result)
                    except Exception as e:
                        self.append_result(f"❌ Error: {file_path}\n   {str(e)}\n\n")

            self.executor = None

            # Display results if not cancelled
            if not self.scan_cancelled:
                self.display_results(results, file_count)

            # Re-enable scan button
            self.scan_btn.config(state="normal")
            self.cancel_btn.config(state="disabled")

        except Exception as e:
            self.append_result(f"\n❌ CRITICAL ERROR: {str(e)}\n")
            self.update_status("Scan failed!", self.accent)
            self.scan_btn.config(state="normal")
            self.cancel_btn.config(state="disabled")

    def display_results(self, results, file_count):
        self.results_text.delete(1.0, tk.END)

        # Count total matches per target
        target_counts = {}
        for r in results:
            for target in r['found']:
                target_counts[target] = target_counts.get(target, 0) + 1

        header = f"{'='*60}\n"
        header += f"★ SCAN COMPLETE ★\n"
        header += f"{'='*60}\n\n"
        header += f"Files scanned: {file_count}\n"
        header += f"Files with matches: {len(results)}\n"

        if target_counts:
            header += f"\nMatches by target:\n"
            for target, count in sorted(target_counts.items()):
                header += f"  • {target}: {count} file(s)\n"

        header += f"\n{'='*60}\n\n"

        self.append_result(header)

        if results:
            for i, r in enumerate(results, 1):
                result_text = f"#{i} 📄 {r['file']}\n"
                result_text += f"   ➤ Found: {', '.join(sorted(r['found']))}\n\n"
                self.append_result(result_text)

            self.update_status(f"Scan complete! Found {len(results)} files with matches", self.bg_secondary)
        else:
            self.append_result("No matches found.\n")
            self.update_status("Scan complete - No matches", self.bg_secondary)

    def append_result(self, text):
        self.root.after(0, lambda: self.results_text.insert(tk.END, text))
        self.root.after(0, lambda: self.results_text.see(tk.END))

    def update_status(self, text, bg=None):
        if bg is None:
            bg = self.bg_secondary
        self.root.after(0, lambda: self.status_bar.config(text=text, bg=bg))

if __name__ == "__main__":
    # Required for multiprocessing on Windows
    multiprocessing.freeze_support()

    root = tk.Tk()
    app = TDScannerGUI(root)
    root.mainloop()